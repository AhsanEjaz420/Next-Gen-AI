import json
import io
import base64
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from app.core.llm import generate_text
from app.prompts.paid.social_calendar_prompt import SOCIAL_CALENDAR_SYSTEM_PROMPT
from app.schemas.paid.social_calendar import SocialCalendarRequest, DailyPost
from pydantic import BaseModel

class SocialCalendar(BaseModel):
    posts: List[DailyPost]

async def generate_social_calendar(data: SocialCalendarRequest, profile=None) -> Dict[str, Any]:
    """
    Generates 30-day social media calendar using Structured Outputs and Brand Context.
    """
    # Default to current month if not provided
    month = data.month or datetime.utcnow().strftime('%B %Y')
    
    # Smart Fallback Logic
    business_niche = data.business_niche or (profile.business_niche if profile else "General Business")
    goal = data.goal or "General brand awareness and engagement"
    tone = data.tone or (profile.brand_tone if profile else "Professional")
    target_audience = data.target_audience or (profile.target_audience if profile else "General audience")

    user_prompt = f"""
Business Niche: {business_niche}
Goal: {goal}
Platforms: {", ".join(data.platforms)}
Tone: {tone}
Target Audience: {target_audience}
Month: {month}

Brand Website (optional): {data.business_website or "Not provided"}
LinkedIn Page (optional): {data.linkedin_url or "Not provided"}
Instagram Page (optional): {data.instagram_url or "Not provided"}
"""
    
    # Brand context for system prompt
    brand_context = ""
    if profile and profile.brand_description:
        brand_context = f"\nBRAND DESCRIPTION:\n{profile.brand_description}"

    full_system_prompt = SOCIAL_CALENDAR_SYSTEM_PROMPT
    if brand_context:
        full_system_prompt += f"\n\nADDITIONAL CONTEXT:{brand_context}"

    # Use structured generation
    calendar_obj = await generate_text(
        system_prompt=full_system_prompt,
        user_prompt=user_prompt,
        temperature=0.75,
        response_model=SocialCalendar
    )

    calendar_data = [post.dict() for post in calendar_obj.posts]
    
    # Generate files (PDF/Excel generation logic remains the same)
    pdf_bytes = _generate_pdf(calendar_data, month, data.business_niche)
    excel_bytes = _generate_excel(calendar_data, month, data.business_niche)
    
    # Encode to base64
    pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
    excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
    
    return {
        "calendar": calendar_data,
        "pdf_base64": pdf_base64,
        "excel_base64": excel_base64,
        "month": month,
        "business_niche": data.business_niche
    }


# ==============================
# PDF GENERATION
# ==============================
def _generate_pdf(calendar: List[dict], month: str, business_niche: str) -> bytes:
    """
    Generate a professional PDF with proper formatting and text wrapping
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        leftMargin=20,
        rightMargin=20,
        topMargin=30,
        bottomMargin=30
    )
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_text = f"<b>30-Day Social Media Content Calendar</b><br/>{month}<br/><font size=10>{business_niche}</font>"
    elements.append(Paragraph(title_text, styles["Title"]))
    elements.append(Paragraph("<br/>", styles["Normal"]))
    
    # Table headers
    table_data = [
        ["Day", "Platform", "Title", "Description", "Type", "CTA"]
    ]
    
    # Add data with Paragraphs for text wrapping
    for item in calendar:
        table_data.append([
            str(item.get("day", "")),
            str(item.get("platform", "")),
            Paragraph(str(item.get("title", "")), styles["Normal"]),
            Paragraph(str(item.get("description", "")), styles["Normal"]),
            Paragraph(str(item.get("content_type", "")), styles["Normal"]),
            Paragraph(str(item.get("cta", "")), styles["Normal"]),
        ])
    
    # Define column widths (in points, total ~550 for A4 with margins)
    col_widths = [30, 70, 100, 150, 80, 120]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header styling
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A90E2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        
        # Body styling
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (1, -1), "CENTER"),  # Center Day & Platform
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ("TOPPADDING", (0, 1), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Footer
    footer_text = f"<br/><br/><font size=8>Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</font>"
    elements.append(Paragraph(footer_text, styles["Normal"]))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


# ==============================
# EXCEL GENERATION
# ==============================
def _generate_excel(calendar: List[dict], month: str, business_niche: str) -> bytes:
    """
    Generate Excel file with formatted data using openpyxl
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(calendar)

    # Reorder columns for better readability
    column_order = ["day", "platform", "title", "description", "content_type", "cta"]
    df = df[[col for col in column_order if col in df.columns]]

    output = io.BytesIO()

    # Write dataframe using openpyxl engine
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="30-Day Calendar", startrow=3)

        workbook = writer.book
        worksheet = writer.sheets["30-Day Calendar"]

        # Styles
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        wrap_align = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title rows
        worksheet["A1"] = f"Social Media Calendar - {month}"
        worksheet["A1"].font = Font(bold=True, size=14)

        worksheet["A2"] = f"Business: {business_niche}"
        worksheet["A2"].font = Font(italic=True, size=10)

        # Style header row
        header_row = 4
        for col_num, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=header_row, column=col_num, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # Style data rows
        for row in range(header_row + 1, header_row + 1 + len(df)):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = wrap_align
                cell.border = thin_border

        # Column widths
        widths = [8, 12, 25, 40, 15, 30]
        for i, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

    output.seek(0)
    return output.read()
